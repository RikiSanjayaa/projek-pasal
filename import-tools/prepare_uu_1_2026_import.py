from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PDF_PATH = ROOT / "uu-nomor-1-tahun-2026.pdf"
OUTPUT_DIR = ROOT / "data-import" / "uu-1-2026"


UU_RECORD = {
    "kode": "UU_1_2026",
    "nama": "UU Penyesuaian Pidana",
    "nama_lengkap": "Undang-Undang Republik Indonesia Nomor 1 Tahun 2026 tentang Penyesuaian Pidana",
    "deskripsi": "Undang-Undang tentang penyesuaian pidana, termasuk perubahan ketentuan pidana dan lampiran penyesuaian.",
    "tahun": 2026,
}


ROMAN_TITLES = {
    "I": "Penyesuaian pidana minimum khusus",
    "II": "Penyesuaian pidana di luar KUHP",
    "III": "Penyesuaian pidana penjara tertentu",
    "IV": "Penyesuaian pidana dalam Peraturan Daerah",
    "V": "Perubahan UU Pembentukan Peraturan Perundang-undangan",
    "VI": "Perubahan UU Pemerintahan Daerah",
    "VII": "Perubahan UU Nomor 1 Tahun 2023 tentang KUHP",
    "VIII": "Ketentuan penyesuaian yang belum tercantum dalam lampiran",
    "IX": "Ketentuan mulai berlaku",
}


CHANGE_SOURCE_BY_ROMAN = {
    "V": "Perubahan UU Nomor 12 Tahun 2011 tentang Pembentukan Peraturan Perundang-undangan",
    "VI": "Perubahan UU Nomor 23 Tahun 2014 tentang Pemerintahan Daerah",
    "VII": "Perubahan UU Nomor 1 Tahun 2023 tentang Kitab Undang-Undang Hukum Pidana",
}


@dataclass
class Marker:
    kind: str
    nomor: str
    start: int
    end: int


def normalize_common_ocr(text: str) -> str:
    text = text.replace("\r", "")
    text = re.sub(r"Undang-\s+Undang", "Undang-Undang", text)
    text = re.sub(r"(?<=\d)[Oo](?=\d|\b)", "0", text)
    text = re.sub(r"(?<=\d)[Il](?=\d|\b)", "1", text)
    text = re.sub(r"\((?:l|I)\)", "(1)", text)
    text = re.sub(r"\bTahun\s+2O", "Tahun 20", text)
    text = re.sub(r"\bNomor\s+l\b", "Nomor 1", text)
    text = text.replace("PEI{YESUAIAN", "PENYESUAIAN")
    text = text.replace("PET{YESUAIAN", "PENYESUAIAN")
    text = text.replace("PENYESUAIAN PIDANA.", "PENYESUAIAN PIDANA.")
    replacements = {
        "Ketenluan": "Ketentuan",
        "Ketenaganukliran": "Ketenaganukliran",
        "lentang": "tentang",
        "tenlang": "tentang",
        "tenta:rg": "tentang",
        "Tahtn": "Tahun",
        "Tahun2022": "Tahun 2022",
        "Undang- Undang": "Undang-Undang",
        "dan/ atau": "dan/atau",
        "dan/atau": "dan/atau",
        "tqiuh": "tujuh",
        "tqjuh": "tujuh",
        "tqiuan": "tujuan",
        "tqjuan": "tujuan",
        "ditqjukan": "ditujukan",
        "ter,.tang": "tentang",
        "Tallun": "Tahun",
        "Ayar (21": "Ayat (2)",
        "l0%": "10%",
        "l0 ": "10 ",
        "l0(": "10(",
        "PTIESIDEN": "PRESIDEN",
        "FRESIDEN": "PRESIDEN",
        "REPUBUK": "REPUBLIK",
        "REPUELIK": "REPUBLIK",
        "REPUEUK": "REPUBLIK",
        "REFUBLIK": "REPUBLIK",
        "tindak pidana": "Tindak Pidana",
    }
    for wrong, right in replacements.items():
        text = text.replace(wrong, right)
    return text


def normalize_nomor(nomor: str) -> str:
    nomor = nomor.strip()
    nomor = nomor.replace("O", "0").replace("o", "0")
    nomor = re.sub(r"(?<=\d)[lI](?=\d|\b)", "1", nomor)
    return nomor


def clean_line(line: str) -> str:
    line = re.sub(r"\s+", " ", line).strip()
    if not line:
        return ""
    if re.match(r"^Pasal\b", line, re.I):
        return normalize_common_ocr(line)
    if re.fullmatch(r"-\s*\d+\s*-?", line):
        return ""
    if re.match(r"^SK\s*No", line, re.I):
        return ""
    if re.match(r"^(PRESIDEN|REPUBLIK|REFUBLIK|REPUELIK|REPUEUK|REPUBUK|BUK|INDONESIA)$", line, re.I):
        return ""
    if re.search(r"EIE|FIT|FII|trN|ItrN", line) and len(line) < 34:
        return ""
    return normalize_common_ocr(line)


def clean_body_text(text: str) -> str:
    lines = []
    for raw in text.splitlines():
        line = clean_line(raw)
        if line:
            lines.append(line)
    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_legal_text(text: str) -> str:
    text = normalize_common_ocr(text)
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    lines = [line for line in lines if line]

    blocks: list[str] = []
    starts_new = re.compile(
        r"^(BAB|Bagian|Paragraf|Pasal\s+|Ayat\s+|\([0-9ivxlcdm]+\)|[a-z]\.|[0-9]+\.|Huruf\s+[a-z]\b)",
        re.I,
    )
    for line in lines:
        if not blocks or starts_new.match(line):
            blocks.append(line)
        else:
            blocks[-1] = f"{blocks[-1]} {line}".strip()

    return "\n".join(blocks).strip()


def strip_next_amendment_intro(text: str) -> str:
    kept = []
    for line in text.splitlines():
        normalized = normalize_common_ocr(line)
        if re.match(r"^\d+\.\s+Ketentuan\b", normalized, re.I):
            break
        kept.append(line)

    return "\n".join(kept).strip()


def extract_pages_text(doc: fitz.Document, first_page: int, last_page: int) -> str:
    raw = []
    for page_number in range(first_page, last_page + 1):
        raw.append(doc.load_page(page_number - 1).get_text("text"))
    return clean_body_text("\n".join(raw))


def find_markers(text: str) -> list[Marker]:
    markers: list[Marker] = []
    roman_re = re.compile(r"(?m)^Pasal\s+([IVXLCDM]+)$")
    numeric_re = re.compile(r"(?m)^Pasal\s+([0-9][0-9A-Za-zOIl]*)$")

    for match in roman_re.finditer(text):
        markers.append(Marker("roman", match.group(1), match.start(), match.end()))
    for match in numeric_re.finditer(text):
        markers.append(Marker("numeric", normalize_nomor(match.group(1)), match.start(), match.end()))

    return sorted(markers, key=lambda marker: marker.start)


def segment_until_next_marker(text: str, markers: list[Marker], index: int) -> str:
    marker = markers[index]
    next_start = markers[index + 1].start if index + 1 < len(markers) else len(text)
    return text[marker.end:next_start].strip()


def extract_explanations(doc: fitz.Document) -> dict[str, str]:
    explanation_text = extract_pages_text(doc, 54, 71)
    markers = [m for m in find_markers(explanation_text) if m.kind == "roman"]
    result: dict[str, str] = {}

    for index, marker in enumerate(markers):
        next_start = markers[index + 1].start if index + 1 < len(markers) else len(explanation_text)
        content = explanation_text[marker.end:next_start].strip()
        result[marker.nomor] = compact_legal_text(content)

    return result


def extract_main_pasal_rows(doc: fitz.Document) -> list[dict[str, str]]:
    body_text = extract_pages_text(doc, 1, 51)
    markers = find_markers(body_text)
    explanations = extract_explanations(doc)
    rows = []

    for index, marker in enumerate(markers):
        if marker.kind != "roman":
            continue
        content = segment_until_next_marker(body_text, markers, index)
        content = re.sub(r"\n?PRESIDEN REPUBLIK INDONESIA,.*$", "", content, flags=re.I | re.S).strip()
        title = ROMAN_TITLES.get(marker.nomor, "Pasal utama UU Nomor 1 Tahun 2026")
        rows.append(
            {
                "nomor": marker.nomor,
                "judul": title,
                "isi": compact_legal_text(content),
                "penjelasan": explanations.get(marker.nomor, ""),
                "keywords": "UU 1 2026, penyesuaian pidana, pasal utama",
            }
        )

    return rows


def nearest_roman_before(markers: list[Marker], position: int) -> str | None:
    roman = None
    for marker in markers:
        if marker.start >= position:
            break
        if marker.kind == "roman":
            roman = marker.nomor
    return roman


def extract_changed_pasal_rows(doc: fitz.Document) -> list[dict[str, str]]:
    body_text = extract_pages_text(doc, 1, 51)
    markers = find_markers(body_text)
    rows = []

    for index, marker in enumerate(markers):
        if marker.kind != "numeric":
            continue

        roman_parent = nearest_roman_before(markers, marker.start)
        if roman_parent not in CHANGE_SOURCE_BY_ROMAN:
            continue

        content = strip_next_amendment_intro(segment_until_next_marker(body_text, markers, index))
        if not content:
            continue

        source = CHANGE_SOURCE_BY_ROMAN[roman_parent]
        rows.append(
            {
                "nomor": marker.nomor,
                "judul": f"Perubahan ketentuan Pasal {marker.nomor}",
                "isi": compact_legal_text(content),
                "penjelasan": (
                    "Data ini bukan pasal utama UU Nomor 1 Tahun 2026. "
                    f"Ini adalah rumusan pasal aturan lain yang diubah melalui {source}."
                ),
                "keywords": "UU 1 2026, perubahan pasal, penyesuaian pidana",
            }
        )

    seen = set()
    unique_rows = []
    for row in rows:
        key = row["nomor"]
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)

    return unique_rows


def write_xlsx(path: Path, rows: list[dict[str, str]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "import"

    headers = ["nomor", "judul", "isi", "penjelasan", "keywords"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    widths = {
        "A": 18,
        "B": 48,
        "C": 90,
        "D": 70,
        "E": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    meta = wb.create_sheet("catatan")
    meta.append(["Judul", title])
    meta.append(["Jumlah baris", len(rows)])
    meta.append(["Sumber PDF", PDF_PATH.name])
    meta.append(["Catatan", "Upload sheet pertama bernama 'import' lewat menu Bulk Import. Pilih UU Nomor 1 Tahun 2026 sebelum import."])
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_readme(output_dir: Path, main_count: int, changed_count: int) -> None:
    readme = f"""# Import UU Nomor 1 Tahun 2026

Folder ini berisi file Excel yang disiapkan dari `uu-nomor-1-tahun-2026.pdf`.

## Data Undang-Undang yang perlu dibuat dulu di Admin

- Kode: `{UU_RECORD['kode']}`
- Nama: `{UU_RECORD['nama']}`
- Nama lengkap: `{UU_RECORD['nama_lengkap']}`
- Tahun: `{UU_RECORD['tahun']}`
- Deskripsi: `{UU_RECORD['deskripsi']}`

## File Import

1. `01_pasal_utama_uu_1_2026_ready.xlsx`
   - Isi: {main_count} pasal utama UU Nomor 1 Tahun 2026.
   - Ini file paling aman untuk import pertama.

2. `02_pasal_perubahan_batang_tubuh_ready.xlsx`
   - Isi: {changed_count} pasal aturan lain yang diubah di batang tubuh UU Nomor 1 Tahun 2026.
   - Tetap sah sebagai isi hukum, tapi konteksnya adalah pasal perubahan.

## Cara Import

1. Buka admin dashboard.
2. Masuk menu `Undang-Undang`.
3. Buat data UU dengan informasi di atas.
4. Masuk menu `Import Data`.
5. Pilih UU Nomor 1 Tahun 2026.
6. Upload file `01_pasal_utama_uu_1_2026_ready.xlsx`.
7. Setelah berhasil, upload `02_pasal_perubahan_batang_tubuh_ready.xlsx`.

## Catatan

Lampiran I dan Lampiran II belum dimasukkan ke file ready karena bentuknya tabel dan ekstraksi PDF masih perlu review manual. Jangan import lampiran mentah sebelum datanya dicek.
"""
    (output_dir / "README_IMPORT.md").write_text(readme, encoding="utf-8")


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(f"PDF tidak ditemukan: {PDF_PATH}")

    doc = fitz.open(PDF_PATH)
    main_rows = extract_main_pasal_rows(doc)
    changed_rows = extract_changed_pasal_rows(doc)

    write_xlsx(
        OUTPUT_DIR / "01_pasal_utama_uu_1_2026_ready.xlsx",
        main_rows,
        "Pasal utama UU Nomor 1 Tahun 2026",
    )
    write_xlsx(
        OUTPUT_DIR / "02_pasal_perubahan_batang_tubuh_ready.xlsx",
        changed_rows,
        "Pasal perubahan di batang tubuh UU Nomor 1 Tahun 2026",
    )
    write_readme(OUTPUT_DIR, len(main_rows), len(changed_rows))

    print(f"Output: {OUTPUT_DIR}")
    print(f"Pasal utama: {len(main_rows)}")
    print(f"Pasal perubahan batang tubuh: {len(changed_rows)}")


if __name__ == "__main__":
    main()
